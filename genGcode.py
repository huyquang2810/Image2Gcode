import insightface
import cv2
import numpy as np
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import rembg
from PIL import Image
import os
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from sklearn.cluster import DBSCAN
from scipy.spatial.distance import cdist

# Adaptive Resampling (góc + khoảng cách)
def simplify_and_adaptive_resample(points, simplify_epsilon=1.0, angle_thresh=10, min_spacing=4):
    if len(points) < 3:
        return points
    # Douglas-Peucker simplification cv2.approxPolyDP
    approx = cv2.approxPolyDP(points, epsilon=simplify_epsilon, closed=True)
    if len(approx) < 3:
        approx = points

    approx = approx.squeeze()
    keep_points = [approx[0]]

    for i in range(1, len(approx) - 1):
        p_prev = approx[i - 1]
        p_curr = approx[i]
        p_next = approx[i + 1]

        v1 = p_curr - p_prev
        v2 = p_next - p_curr

        cos_angle = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2) + 1e-8)
        angle = np.degrees(np.arccos(np.clip(cos_angle, -1.0, 1.0)))

        dist = np.linalg.norm(p_curr - keep_points[-1])
        if angle < (180 - angle_thresh) or dist >= min_spacing:
            keep_points.append(p_curr)

    keep_points.append(approx[-1])
    return np.array(keep_points, dtype=np.int32).reshape(-1, 1, 2)

# --- Class xử lý ảnh và sinh G-code ---
class Picture:
    def __init__(self, filepath, x_max=150, y_max=150):
        self.img = Image.open(filepath).convert("RGB")
        self.img = np.array(self.img)
        self.h, self.w, self.c = self.img.shape
        self.pre = np.ones(self.img.shape)
        self.gcode = ['G28']
        self.x_max = x_max
        self.y_max = y_max

    def gray_scale(self):
        gray = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
        self.gray = gray / 255.0  # For visualization

        blurred = cv2.GaussianBlur(gray, (3, 3), 0)
        edges = cv2.Canny(blurred, threshold1=50, threshold2=150)

        binary = (edges / 255.0).astype(float)
        # binary = 1.0 - (edges / 255.0).astype(float)  # <- đảo ngược để có nền trắng
        self.pre = np.stack([binary] * 3, axis=-1)
        return self.pre


    # Ap dụng thử opening/ closing nhưng chưa cho ra kết quả tốt
    # def gray_scale(self):
    #     gray = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
    #     self.gray = gray / 255.0
    #
    #     # Làm mượt vừa đủ
    #     blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    #
    #     # Mở để bỏ nhiễu, Closing để nối lại các nét quan trọng
    #     kernel = np.ones((2, 2), np.uint8)
    #     opened = cv2.morphologyEx(blurred, cv2.MORPH_OPEN, kernel)
    #     closed = cv2.morphologyEx(opened, cv2.MORPH_CLOSE, kernel)
    #
    #     # Canny với ngưỡng cao hơn → bớt nét mảnh vụn
    #     edges = cv2.Canny((closed * 255).astype(np.uint8), threshold1=50, threshold2=140)
    #
    #     binary = (edges / 255.0).astype(float)
    #     self.pre = np.stack([binary] * 3, axis=-1)
    #     return self.pre

    # def gray_scale(self):
    #     gray = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
    #     self.gray = gray / 255.0
    #
    #     # Làm sạch noise trước khi edge detect
    #     blurred = cv2.GaussianBlur(gray, (3, 3), 0)  # Blur mạnh hơn (5x5)
    #
    #     # Morphological opening để bỏ noise nhỏ
    #     kernel = np.ones((3, 3), np.uint8)
    #     opened = cv2.morphologyEx(blurred, cv2.MORPH_OPEN, kernel)
    #
    #     edges = cv2.Canny((opened * 255).astype(np.uint8), threshold1=50, threshold2=150)
    #     binary = (edges / 255.0).astype(float)
    #
    #     self.pre = np.stack([binary] * 3, axis=-1)
    #     return self.pre

    def save_gray(self, output):
        if hasattr(self, 'gray') and self.gray is not None:
            plt.imshow(self.gray, cmap='gray')
            plt.axis('off')
            plt.imsave(output + '_gray.jpg', self.gray, cmap='gray')
            # plt.imsave(os.path.normpath(output + '_gray.jpg'), self.gray, cmap='gray')
            print('✅ Saved ' + output + '_gray.jpg')

    def save_binary(self, output):

        if hasattr(self, 'pre') and self.pre is not None:
            # Đảo màu: nét đen (0), nền trắng (255)
            binary_inverted = (1 - self.pre)  # đảo ngược: trắng → đen, đen → trắng
            plt.imshow(binary_inverted, cmap='gray')
            plt.axis('off')
            plt.imsave(output + '_binary.jpg', binary_inverted, cmap='gray')
            print('✅ Saved ' + output + '_binary.jpg ')

        # plt.imshow(self.pre, cmap='gray')
        # plt.axis('off')
        # plt.imsave(output + '_binary.jpg', self.pre)
        # print('✅ Saved ' + output + '_binary.jpg')
    # hàm gốc chưa chỉnh
    # def gen_gcode(self):
    #     binary = (self.pre[:, :, 0] > 0.5).astype(np.uint8) * 255
    #     contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    #     ratio = self.x_max / max(self.w, self.h)
    #     total_points = 0
    #
    #     spindle_on = False  # Trạng thái của spindle (M3)
    #
    #     for contour in contours:
    #         if len(contour) < 2:
    #             continue
    #         total_points += len(contour)
    #
    #         # Di chuyển đến điểm đầu tiên (G0 + M5 + G92)
    #         x0, y0 = contour[0][0]
    #         y0_flipped = self.h - y0
    #         if spindle_on:
    #             self.gcode.append("M5")  # Tắt spindle trước khi di chuyển
    #             spindle_on = False
    #         self.gcode.append(f"G0 X{x0 * ratio:.4f} Y{y0_flipped * ratio:.4f}")  # Di chuyển nhanh
    #         self.gcode.append(f"G92 X{x0 * ratio:.4f} Y{y0_flipped * ratio:.4f}")  # Thiết lập vị trí
    #
    #         # Bắt đầu vẽ (G1 + M3 + G92)
    #         for pt in contour[1:]:
    #             x, y = pt[0]
    #             y_flipped = self.h - y
    #             if not spindle_on:
    #                 self.gcode.append("M3")  # Bật spindle trước khi vẽ
    #                 spindle_on = True
    #             self.gcode.append(f"G1 X{x * ratio:.2f} Y{y_flipped * ratio:.2f}")  # Vẽ
    #             self.gcode.append(f"G92 X{x * ratio:.2f} Y{y_flipped * ratio:.2f}")  # Cập nhật vị trí
    #
    #     # Kết thúc nếu spindle đang bật thì tắt đi (M5)
    #     if spindle_on:
    #         self.gcode.append("M5")
    #         spindle_on = False
    #
    #     return self.gcode, total_points

    # eps=1, simplify_epsilon=1, 10,1,4,10/8,1.2,6,12,/ 10,0.1,1,10
    # def gen_gcode(self, eps= 8, simplify_epsilon=1.2, min_spacing=6, min_contour_len=12):
    #     binary = (self.pre[:, :, 0] > 0.5).astype(np.uint8) * 255
    #     contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    #     ratio = self.x_max / max(self.w, self.h)
    #     total_points = 0
    #
    #     # Tính center mỗi contour
    #     centers = []
    #     valid_contours = []
    #     for contour in contours:
    #         if len(contour) < min_contour_len:
    #             continue
    #         M = cv2.moments(contour)
    #         if M["m00"] != 0:
    #             cx = int(M["m10"] / M["m00"])
    #             cy = int(M["m01"] / M["m00"])
    #         else:
    #             cx, cy = contour[0][0]
    #         centers.append([cx, cy])
    #         valid_contours.append(contour)
    #
    #     # Gom cụm bằng DBSCAN
    #     if not centers:
    #         return self.gcode, total_points
    #     labels = DBSCAN(eps=eps, min_samples=1).fit_predict(centers)
    #
    #     clusters = {}
    #     for label, contour in zip(labels, valid_contours):
    #         clusters.setdefault(label, []).append(contour)
    #
    #     for cluster in clusters.values():
    #         for contour in cluster:  # ❗ Xử lý từng contour riêng, không nối lại
    #             simplified = simplify_and_adaptive_resample(
    #                 contour,
    #                 simplify_epsilon=simplify_epsilon,
    #                 angle_thresh=15,
    #                 min_spacing=min_spacing
    #             )
    #             if len(simplified) < 2:
    #                 continue
    #             total_points += len(simplified)
    #             x0, y0 = simplified[0][0]
    #             y0_flipped = self.h - y0
    #             self.gcode.append(f"G0 X{x0 * ratio:.2f} Y{y0_flipped * ratio:.2f}")
    #             for pt in simplified[1:]:
    #                 x, y = pt[0]
    #                 y_flipped = self.h - y
    #                 self.gcode.append(f"G1 X{x * ratio:.2f} Y{y_flipped * ratio:.2f}")
    #
    #     return self.gcode, total_points

    # eps=1, simplify_epsilon=1, 10,1,4,10/8,1.2,6,12,/ 10,0.1,1,10
    def gen_gcode(self, eps=10, simplify_epsilon=1, min_spacing=4, min_contour_len=10):
        binary = (self.pre[:, :, 0] > 0.5).astype(np.uint8) * 255
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        ratio = self.x_max / max(self.w, self.h)
        total_points = 0

        centers = []
        valid_contours = []
        for contour in contours:
            if len(contour) < min_contour_len:
                continue
            M = cv2.moments(contour)
            if M["m00"] != 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])
            else:
                cx, cy = contour[0][0]
            centers.append([cx, cy])
            valid_contours.append(contour)

        if not centers:
            return self.gcode, total_points

        labels = DBSCAN(eps=eps, min_samples=1).fit_predict(centers)

        clusters = {}
        for label, contour in zip(labels, valid_contours):
            clusters.setdefault(label, []).append(contour)

        spindle_on = False  # trạng thái spindle

        for cluster in clusters.values():
            for contour in cluster:
                simplified = simplify_and_adaptive_resample(
                    contour,
                    simplify_epsilon=simplify_epsilon,
                    angle_thresh=15,
                    min_spacing=min_spacing
                )
                if len(simplified) < 2:
                    continue
                total_points += len(simplified)

                # Di chuyển nhanh đến điểm đầu (G0 + M5 + G92)
                x0, y0 = simplified[0][0]
                y0_flipped = self.h - y0
                if spindle_on:
                    self.gcode.append("M5")  # Tắt spindle trước di chuyển nhanh
                    spindle_on = False
                self.gcode.append(f"G0 X{x0 * ratio:.2f} Y{y0_flipped * ratio:.2f}")
                self.gcode.append(f"G92 X{x0 * ratio:.2f} Y{y0_flipped * ratio:.2f}")

                # Vẽ (G1 + M3 + G92)
                for pt in simplified[1:]:
                    x, y = pt[0]
                    y_flipped = self.h - y
                    if not spindle_on:
                        self.gcode.append("M3")  # Bật spindle trước khi vẽ
                        spindle_on = True
                    self.gcode.append(f"G1 X{x * ratio:.2f} Y{y_flipped * ratio:.2f}")
                    self.gcode.append(f"G92 X{x * ratio:.2f} Y{y_flipped * ratio:.2f}")

        # Kết thúc nếu spindle đang bật thì tắt đi
        if spindle_on:
            self.gcode.append("M5")
            spindle_on = False

        return self.gcode, total_points

    #TSP
    # def gen_gcode(self, eps=8, simplify_epsilon=1.2, min_spacing=6, min_contour_len=12):
    #     binary = (self.pre[:, :, 0] > 0.5).astype(np.uint8) * 255
    #     contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    #     ratio = self.x_max / max(self.w, self.h)
    #     total_points = 0
    #
    #     # Lọc và lấy center của các contour đủ lớn
    #     centers = []
    #     filtered = []
    #     for contour in contours:
    #         if len(contour) < min_contour_len:
    #             continue
    #         M = cv2.moments(contour)
    #         if M["m00"] != 0:
    #             cx = int(M["m10"] / M["m00"])
    #             cy = int(M["m01"] / M["m00"])
    #         else:
    #             cx, cy = contour[0][0]
    #         centers.append([cx, cy])
    #         filtered.append(contour)
    #
    #     if not centers:
    #         return self.gcode, total_points
    #
    #     # TSP: sắp xếp thứ tự contour theo đường đi ngắn nhất (Nearest Neighbor)
    #     visited = set()
    #     order = []
    #     current = 0
    #     visited.add(current)
    #     order.append(current)
    #     while len(order) < len(centers):
    #         dists = cdist([centers[current]], centers)[0]
    #         dists[list(visited)] = np.inf
    #         next_idx = np.argmin(dists)
    #         visited.add(next_idx)
    #         order.append(next_idx)
    #         current = next_idx
    #
    #     sorted_contours = [filtered[i] for i in order]
    #
    #     # Sinh G-code từ contour đã sắp xếp
    #     for contour in sorted_contours:
    #         simplified = simplify_and_adaptive_resample(
    #             contour,
    #             simplify_epsilon=simplify_epsilon,
    #             angle_thresh=15,
    #             min_spacing=min_spacing
    #         )
    #         if len(simplified) < 2:
    #             continue
    #         total_points += len(simplified)
    #         x0, y0 = simplified[0][0]
    #         y0_flipped = self.h - y0
    #         self.gcode.append(f"G0 X{x0 * ratio:.2f} Y{y0_flipped * ratio:.2f}")
    #         for pt in simplified[1:]:
    #             x, y = pt[0]
    #             y_flipped = self.h - y
    #             self.gcode.append(f"G1 X{x * ratio:.2f} Y{y_flipped * ratio:.2f}")
    #
    #     return self.gcode, total_points

    def save_gcode(self, output_name):
        os.makedirs(os.path.dirname(output_name), exist_ok=True)
        with open(f'{output_name}_gcode.nc', 'w') as f:
            for line in self.gcode:
                f.write(f'{line}\n')
        print(f'✅ Saved {output_name}_gcode.nc')

# --- Tiện ích resize ảnh trước khi gắn vào Excel ---
def resize_and_save_temp(image_path, output_path, max_size=(100, 100)):
    try:
        img = PILImage.open(image_path).convert("RGB")  # ← Thêm convert tại đây
        img.thumbnail(max_size)
        img.save(output_path)
        return True
    except Exception as e:
        print(f"⚠️ Lỗi khi resize ảnh: {image_path} → {e}")
        return False

# --- Xử lý khuôn mặt ---
def init_face_analyzer():
    print("Initializing face analysis...")
    face_analyzer = insightface.app.FaceAnalysis(name='buffalo_l')
    try:
        face_analyzer.prepare(ctx_id=0)
        print(f"✅ Using GPU: {face_analyzer.models['detection'].providers}")
    except Exception:
        print("GPU initialization failed. Switching to CPU...")
        face_analyzer.prepare(ctx_id=-1)
    return face_analyzer

def align_face(image, face):
    left_eye = tuple(face.kps[0].astype(int))
    right_eye = tuple(face.kps[1].astype(int))
    eyes_center = tuple(map(int, ((left_eye[0] + right_eye[0]) / 2, (left_eye[1] + right_eye[1]) / 2)))
    dx = right_eye[0] - left_eye[0]
    dy = right_eye[1] - left_eye[1]
    angle = np.degrees(np.arctan2(dy, dx))
    rotation_matrix = cv2.getRotationMatrix2D(eyes_center, angle, 1)
    aligned = cv2.warpAffine(image.astype(np.float32), rotation_matrix, (image.shape[1], image.shape[0]),
                             flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_CONSTANT, borderValue=(255, 255, 255))
    return np.clip(aligned, 0, 255).astype(np.uint8)

def resize_to_a4(image, target_width=300, target_height=300):
    h, w = image.shape[:2]
    scale = min(target_width / w, target_height / h)
    new_w, new_h = int(w * scale), int(h * scale)
    resized = cv2.resize(image, (new_w, new_h), interpolation=cv2.INTER_AREA)
    canvas = np.ones((target_height, target_width, 3), dtype=np.uint8) * 255
    top = (target_height - new_h) // 2
    left = (target_width - new_w) // 2
    canvas[top:top + new_h, left:left + new_w] = resized
    return canvas

# === Main pipeline ===
if __name__ == '__main__':
    face_model = init_face_analyzer()

    print("Chọn chế độ:\n1: Folder\n2: Webcam\n3: xử lý ảnh cuối cùng trong thư mục")

    mode = input("Nhập chế độ: ")

    if mode == '1':
        input_face_dir = "Input_Folder"
        filenames = sorted(
            [f for f in os.listdir(input_face_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
            key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
        )

    elif mode == '2':
        input_face_dir = "Cam_Input"
        os.makedirs(input_face_dir, exist_ok=True)
        cap = cv2.VideoCapture(0, cv2.CAP_ANY)
        if not cap.isOpened():
            print("❌ Webcam không thể truy cập.")
            exit()
        count = 1
        print("Nhấn Enter để chụp, ESC để thoát.")
        while True:
            ret, frame = cap.read()
            if not ret:
                print("❌ Không lấy được frame.")
                break
            cv2.imshow("Webcam", frame)
            key = cv2.waitKey(1) & 0xFF
            if key == 13:
                save_path = os.path.join(input_face_dir, f"cam_{count}.jpg")
                cv2.imwrite(save_path, frame)
                print(f"📸 Đã chụp: {save_path}")
                count += 1
            elif key == 27:
                break
        cap.release()
        cv2.destroyAllWindows()
        filenames = sorted(
            [f for f in os.listdir(input_face_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
            key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
        )

    elif mode == '3':
        input_face_dir = "Input_Folder"
        files = [f for f in os.listdir(input_face_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        if not files:
            print("❌ Không tìm thấy ảnh trong thư mục.")
            exit()

        # Lấy ảnh có thời gian sửa đổi gần nhất
        latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(input_face_dir, x)))
        filenames = [latest_file]
        print(f"🕓 Ảnh mới nhất được xử lý: {latest_file}")


    else:
        print("Chế độ không hợp lệ.")
        exit()

    output_folder = 'out'
    excel_folder = 'excel'
    thumb_folder = os.path.join(excel_folder, 'thumbs')
    img_output = "img"

    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(thumb_folder, exist_ok=True)
    os.makedirs(excel_folder, exist_ok=True)
    os.makedirs(img_output, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Thời gian xử lý ảnh"
    ws.append(["Tên file", "Input", "Thời gian (s)", "Output", "Số điểm G-code"])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # filenames = sorted(
    #     [f for f in os.listdir(input_face_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
    #     key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
    # )

    for file in filenames:
        print(f"\n📂 Đang xử lý ảnh: {file}")
        start_time = datetime.now()
        input_path = os.path.join(input_face_dir, file)
        image = cv2.imread(input_path)
        if image is None:
            print(f"⚠️ Không đọc được ảnh: {file}")
            continue

        # Tách nền
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(image_rgb)
        removed = rembg.remove(pil_img)
        rgba = removed.convert("RGBA")
        np_img = np.array(rgba)
        white_bg = np.ones_like(np_img[:, :, :3], dtype=np.uint8) * 255
        alpha = np_img[:, :, 3:4] / 255.0
        blend = (np_img[:, :, :3] * alpha + white_bg * (1 - alpha)).astype(np.uint8)

        # Phát hiện mặt
        faces = face_model.get(blend)
        if not faces:
            print(f"❌ Không phát hiện khuôn mặt trong ảnh: {file}")
            continue

        face = faces[0]
        aligned = align_face(blend, face)

        # === CHÈN ĐOẠN NÀY VÀO ĐÂY ===
        bounding_box_dir = "bounding_box"
        os.makedirs(bounding_box_dir, exist_ok=True)

        boxed_img = blend.copy()
        x1, y1, x2, y2 = face.bbox.astype(int)
        cv2.rectangle(boxed_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        for (x, y) in face.kps.astype(int):
            cv2.circle(boxed_img, (x, y), 2, (0, 0, 255), -1)

        bbox_save_path = os.path.join(bounding_box_dir, os.path.splitext(file)[0] + "_bbox.jpg")
        cv2.imwrite(bbox_save_path, cv2.cvtColor(boxed_img, cv2.COLOR_RGB2BGR))
        print(f"🟩 Bounding box image saved: {os.path.basename(bbox_save_path)}")
        # === HẾT ĐOẠN THÊM ===

        x1, y1, x2, y2 = face.bbox.astype(int)
        h, w = aligned.shape[:2]
        top = np.clip(y1 - int(0.45 * (y2 - y1)), 0, h)
        bottom = np.clip(y2 + int(0.05 * (y2 - y1)), 0, h)
        left = np.clip(x1 - int(0.3 * (x2 - x1)), 0, w)
        right = np.clip(x2 + int(0.3 * (x2 - x1)), 0, w)
        crop = aligned[top:bottom, left:right]
        crop_rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)

        # Resize A4
        a4 = resize_to_a4(crop_rgb)
        a4_name = os.path.splitext(file)[0] + "_a4.jpg"
        a4_path = os.path.join(img_output, a4_name)
        cv2.imwrite(a4_path, a4)
        print(f"✅ Aligned and cropped face: {a4_name}")

        # Xử lý nhị phân & G-code
        output_name = os.path.join(output_folder, os.path.splitext(file)[0])
        pic = Picture(a4_path)
        pic.gray_scale()  # ← Truyền face vào đây
        # pic.gray_scale()
        pic.save_gray(output_name)
        pic.save_binary(output_name)
        gcode, num_points = pic.gen_gcode()
        # gcode, num_points = pic.gen_gcode(eps=10, n_waypoints=80, min_contour_len=40)
        pic.save_gcode(output_name)

        duration = (datetime.now() - start_time).total_seconds()
        ws.append([file, "", round(duration, 3), "", num_points])
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 85

        thumb1 = os.path.join(thumb_folder, f'{os.path.splitext(file)[0]}_thumb1.jpg')
        thumb2 = os.path.join(thumb_folder, f'{os.path.splitext(file)[0]}_thumb2.jpg')

        if resize_and_save_temp(input_path, thumb1):
            ws.add_image(XLImage(thumb1), f'B{current_row}')
        if resize_and_save_temp(f'{output_name}_binary.jpg', thumb2):
            ws.add_image(XLImage(thumb2), f'D{current_row}')

    total_time = sum(ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1))
    ws.append(["", "", f"TỔNG: {total_time:.3f} giây", "", ""])
    ws[f"C{ws.max_row}"].font = Font(bold=True)

    wb.save(os.path.join(excel_folder, 'time_processing.xlsx'))
    print("\n🎉 Xử lý hoàn tất toàn bộ ảnh!")

# if __name__ == '__main__':
#     print("Chọn chế độ:\n1: Folder\n2: Webcam\n3: Xử lý ảnh cuối cùng trong thư mục")
#
#     mode = input("Nhập chế độ: ")
#
#     if mode == '1':
#         input_dir = "Input_Folder"
#         filenames = sorted(
#             [f for f in os.listdir(input_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
#             key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
#         )
#
#     elif mode == '2':
#         input_dir = "Cam_Input"
#         os.makedirs(input_dir, exist_ok=True)
#         cap = cv2.VideoCapture(0, cv2.CAP_ANY)
#         if not cap.isOpened():
#             print("❌ Webcam không thể truy cập.")
#             exit()
#         count = 1
#         print("Nhấn Enter để chụp, ESC để thoát.")
#         while True:
#             ret, frame = cap.read()
#             if not ret:
#                 print("❌ Không lấy được frame.")
#                 break
#             cv2.imshow("Webcam", frame)
#             key = cv2.waitKey(1) & 0xFF
#             if key == 13:  # Enter
#                 save_path = os.path.join(input_dir, f"cam_{count}.jpg")
#                 cv2.imwrite(save_path, frame)
#                 print(f"📸 Đã chụp: {save_path}")
#                 count += 1
#             elif key == 27:  # ESC
#                 break
#         cap.release()
#         cv2.destroyAllWindows()
#
#         filenames = sorted(
#             [f for f in os.listdir(input_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
#             key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
#         )
#
#     elif mode == '3':
#         input_dir = "Input_Folder"
#         files = [f for f in os.listdir(input_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
#         if not files:
#             print("❌ Không tìm thấy ảnh trong thư mục.")
#             exit()
#
#         latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(input_dir, x)))
#         filenames = [latest_file]
#         print(f"🕓 Ảnh mới nhất được xử lý: {latest_file}")
#
#     else:
#         print("Chế độ không hợp lệ.")
#         exit()
#
#     # Tạo thư mục đầu ra
#     output_folder = 'out'
#     excel_folder = 'excel'
#     thumb_folder = os.path.join(excel_folder, 'thumbs')
#     img_output = "img_output"
#
#     os.makedirs(output_folder, exist_ok=True)
#     os.makedirs(thumb_folder, exist_ok=True)
#     os.makedirs(excel_folder, exist_ok=True)
#     os.makedirs(img_output, exist_ok=True)
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Thời gian xử lý ảnh"
#     ws.append(["Tên file", "Input", "Thời gian (s)", "Output", "Số điểm G-code"])
#     ws.column_dimensions['A'].width = 25
#     ws.column_dimensions['B'].width = 14
#     ws.column_dimensions['C'].width = 16
#     ws.column_dimensions['D'].width = 14
#     ws.column_dimensions['E'].width = 14
#
#     for file in filenames:
#         print(f"\n📂 Đang xử lý ảnh: {file}")
#         start_time = datetime.now()
#         input_path = os.path.join(input_dir, file)
#         image = cv2.imread(input_path)
#         if image is None:
#             print(f"⚠️ Không đọc được ảnh: {file}")
#             continue
#
#         # Xử lý tách nền, bạn có thể bỏ đoạn này nếu không muốn tách nền
#         image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
#         pil_img = Image.fromarray(image_rgb)
#         removed = rembg.remove(pil_img)
#         rgba = removed.convert("RGBA")
#         np_img = np.array(rgba)
#         white_bg = np.ones_like(np_img[:, :, :3], dtype=np.uint8) * 255
#         alpha = np_img[:, :, 3:4] / 255.0
#         blend = (np_img[:, :, :3] * alpha + white_bg * (1 - alpha)).astype(np.uint8)
#
#         # Resize về kích thước A4
#         a4 = resize_to_a4(blend)
#         a4_name = os.path.splitext(file)[0] + "_a4.jpg"
#         a4_path = os.path.join(img_output, a4_name)
#         cv2.imwrite(a4_path, a4)
#         print(f"✅ Resize to A4 done: {a4_name}")
#
#         # Xử lý ảnh chuyển sang G-code
#         pic = Picture(a4_path)
#         pic.gray_scale()
#         pic.save_gray(os.path.join(output_folder, os.path.splitext(file)[0]))
#         pic.save_binary(os.path.join(output_folder, os.path.splitext(file)[0]))
#         gcode, num_points = pic.gen_gcode()
#         pic.save_gcode(os.path.join(output_folder, os.path.splitext(file)[0]))
#
#         duration = (datetime.now() - start_time).total_seconds()
#         ws.append([file, "", round(duration, 3), "", num_points])
#         current_row = ws.max_row
#         ws.row_dimensions[current_row].height = 85
#
#         thumb1 = os.path.join(thumb_folder, f'{os.path.splitext(file)[0]}_thumb1.jpg')
#         thumb2 = os.path.join(thumb_folder, f'{os.path.splitext(file)[0]}_thumb2.jpg')
#
#         if resize_and_save_temp(input_path, thumb1):
#             ws.add_image(XLImage(thumb1), f'B{current_row}')
#         if resize_and_save_temp(os.path.join(output_folder, f'{os.path.splitext(file)[0]}_binary.jpg'), thumb2):
#             ws.add_image(XLImage(thumb2), f'D{current_row}')
#
#     total_time = sum(ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1))
#     ws.append(["", "", f"TỔNG: {total_time:.3f} giây", "", ""])
#     ws[f"C{ws.max_row}"].font = Font(bold=True)
#
#     wb.save(os.path.join(excel_folder, 'time_processing.xlsx'))
#     print("\n🎉 Xử lý hoàn tất toàn bộ ảnh!")


